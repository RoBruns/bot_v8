import os
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FONT_NAME = 'Calibri'
COR_HEADER  = 'C8A2C8'
COR_SALDO   = 'C6EFCE'
COR_SEM     = 'FCE4D6'
COR_NAO_AUT = 'FFEB9C'
COR_FALHA   = 'FFC7CE'
COR_NEUTRO  = 'FFFFFF'
COR_RES_HDR = '4472C4'
COR_RES_BDY = 'DDEEFF'


def _font(bold=False, color='000000', size=11):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def _border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(cell, value, bold=False, font_color='000000', bg=None, align='left'):
    cell.value = value
    cell.font = _font(bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if bg:
        cell.fill = _fill(bg)
    cell.border = _border()

def _row_color(valor) -> str:
    try:
        float(valor)
        return COR_SALDO
    except (ValueError, TypeError):
        pass
    v = str(valor)
    if v == 'SEM SALDO':      return COR_SEM
    if v == 'NAO AUTORIZADO': return COR_NAO_AUT
    if v == 'FALHA CONSULTA': return COR_FALHA
    return COR_NEUTRO


def export_session(results: List[Dict], session_info: Dict, output_path: str):
    """
    results: list of dicts with keys cpf, valor, motivo, consulted_at
    session_info: dict with base_name, tabela_simulacao, processed, com_saldo, etc.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.sheet_view.showGridLines = False

    has_motivo = any(r.get("motivo") for r in results)
    if has_motivo:
        col_f, col_v = 'F', 'G'
        ws.column_dimensions['C'].width = 52
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 2
        ws.column_dimensions['F'].width = 26
        ws.column_dimensions['G'].width = 26
    else:
        col_f, col_v = 'E', 'F'
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 26
        ws.column_dimensions['F'].width = 26

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16

    _cell(ws['A1'], 'CPF',   bold=True, font_color='FFFFFF', bg=COR_HEADER)
    _cell(ws['B1'], 'Valor', bold=True, font_color='FFFFFF', bg=COR_HEADER)
    if has_motivo:
        _cell(ws['C1'], 'Motivo da Falha', bold=True, font_color='FFFFFF', bg=COR_HEADER)
        _cell(ws['D1'], 'Data e Hora',     bold=True, font_color='FFFFFF', bg=COR_HEADER)
    else:
        _cell(ws['C1'], 'Data e Hora', bold=True, font_color='FFFFFF', bg=COR_HEADER)

    for i, r in enumerate(results, start=2):
        bg = _row_color(r["valor"])
        valor = r["valor"]
        try:
            fval = float(valor)
            display = f"R$ {fval:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except (ValueError, TypeError):
            display = valor
        _cell(ws[f'A{i}'], r["cpf"],   bg=bg)
        _cell(ws[f'B{i}'], display,    bg=bg, align='center')
        if has_motivo:
            _cell(ws[f'C{i}'], r.get("motivo") or '', bg=bg)
            _cell(ws[f'D{i}'], r.get("consulted_at", ''), bg=bg, align='center')
        else:
            _cell(ws[f'C{i}'], r.get("consulted_at", ''), bg=bg, align='center')

    resumo = [
        ("Base processada",     session_info.get("base_name", "-")),
        ("Tabela de simulação", session_info.get("tabela_simulacao", "-")),
        ("CPFs processados",    session_info.get("processed", 0)),
        ("Com saldo",           session_info.get("com_saldo", 0)),
        ("Sem saldo",           session_info.get("sem_saldo", 0)),
        ("Não autorizado",      session_info.get("nao_autorizado", 0)),
        ("CPF inválido",        session_info.get("cpf_invalido", 0)),
        ("Falha de consulta",   session_info.get("falha_consulta", 0)),
    ]
    _cell(ws[f'{col_f}1'], 'Campo', bold=True, font_color='FFFFFF', bg=COR_RES_HDR)
    _cell(ws[f'{col_v}1'], 'Valor', bold=True, font_color='FFFFFF', bg=COR_RES_HDR)
    for i, (campo, valor) in enumerate(resumo, start=2):
        _cell(ws[f'{col_f}{i}'], campo, bg=COR_RES_BDY)
        _cell(ws[f'{col_v}{i}'], valor, bg=COR_RES_BDY, align='center')

    ws.freeze_panes = 'A2'
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)

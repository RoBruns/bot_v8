# Flake8: noqa
"""
Main script for BOT3.0API.
Handles authentication, balance consultation, and simulation logic.
"""

import getpass
import json
import msvcrt
import os
import sys
import time
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
import requests
from dotenv import load_dotenv, set_key
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import logging
from logger_setup import setup_logger

# --- Constants ---
MAX_RETRIES = 25
TIMEOUT_SECONDS = 10

# File Paths
ENV_FILE_PATH = '.env'
CONTADOR_FILE = 'contador.txt'
RESULT_JSON_DIR = 'result_json'
RESULT_JSON_FILE = os.path.join(RESULT_JSON_DIR, 'result.json')
BASE_DIR = 'base'

# URLs
URL_AUTH_TOKEN = 'https://auth.v8sistema.com/oauth/token'
URL_CONSULT_BALANCE = 'https://bff.v8sistema.com/fgts/balance'
URL_SIMULATION = 'https://bff.v8sistema.com/fgts/simulations'
URL_BOT_AUTH_BASE = 'https://web-production-7127e.up.railway.app/users/'

# API Constants
AUTH_AUDIENCE = 'https://bff.v8sistema.com'
AUTH_CLIENT_ID = 'DHWogdaYmEI8n5bwwxPDzulMlSK7dwIn'
AUTH_SCOPE = 'offline_access'
AUTH_GRANT_TYPE = 'password'
PROVIDER_BMS = "bms"

TABELAS = {
    "1": {"nome": "Normal",  "id": "cb563029-ba93-4b53-8d53-4ac145087212"},
    "2": {"nome": "Cometa",  "id": "61c9fb2f-c902-4992-b8f5-b0ee368c45b0"},
}


class ConsultStatus(str, Enum):
    NAO_AUTORIZADO = "NAO AUTORIZADO"
    SEM_SALDO = "SEM SALDO"
    CPF_INVALIDO = "CPF INVÁLIDO"
    FALHA_CONSULTA = "FALHA CONSULTA"
    RETRY = "RETRY"
    TOKEN_EXPIRADO = "TOKEN EXPIRADO"
    OPERACAO_EM_ANDAMENTO = "OPERAÇÃO EM ANDAMENTO"


def get_token(username: str, password: str) -> Optional[str]:
    """Retrieves the access token from the authentication service."""
    data_get_token = {
        'grant_type': AUTH_GRANT_TYPE,
        'username': username,
        'password': password,
        'audience': AUTH_AUDIENCE,
        'scope': AUTH_SCOPE,
        'client_id': AUTH_CLIENT_ID
    }

    try:
        response = requests.post(URL_AUTH_TOKEN, data=data_get_token, timeout=TIMEOUT_SECONDS)
        response.raise_for_status()
        return response.json().get('access_token')
    except requests.RequestException as e:
        error_msg = f'Erro ao obter o token: {e}'
        print(error_msg)
        logging.error(error_msg)
        return None


def create_session() -> requests.Session:
    """Creates a requests Session with retry logic."""
    session = requests.Session()
    retries = Retry(
        total=MAX_RETRIES,
        backoff_factor=1,
        status_forcelist=[500, 502, 503, 504]
    )
    session.mount('https://', HTTPAdapter(max_retries=retries))
    return session


def _handle_consult_error(response: requests.Response, cpf: str) -> Tuple[Any, Any, bool, Optional[str]]:
    """Handles errors during balance consultation.
    Returns: (status, balance_id, finished, error_reason)
    error_reason is only set when status is FALHA_CONSULTA.
    """
    try:
        raw = response.text
        try:
            body = response.json()
        except ValueError:
            # Resposta não-JSON (ex: timeout de proxy/load balancer)
            logging.warning(f"CPF {cpf}: resposta não-JSON (status {response.status_code}): {raw[:200]}")
            return ConsultStatus.RETRY, None, False, None
        error_msg = str(body.get('error', ''))

        if response.status_code == 400:
            detail = body.get('detail', '')
            combined_msg = detail + " " + error_msg
            print(f"Erro 400 para o CPF {cpf}: {combined_msg.strip()}")
            
            if "fiduciária em andamento" in combined_msg.lower() or "fiduciaria em andamento" in combined_msg.lower() or "fiduciaria em adamento" in combined_msg.lower():
                logging.info(f"CPF {cpf}: Operação Fiduciária em andamento (400) - Detalhe: {combined_msg.strip()}")
                return ConsultStatus.OPERACAO_EM_ANDAMENTO, None, True, None
            if "não possui autorização" in combined_msg or "Fiduciária" in combined_msg:
                logging.info(f"CPF {cpf}: Não autorizado (400) - Detalhe: {combined_msg.strip()}")
                return ConsultStatus.NAO_AUTORIZADO, None, True, None
            if "não possui saldo disponível" in combined_msg.lower() or "saldo insuficiente" in combined_msg.lower():
                logging.info(f"CPF {cpf}: Sem saldo (400)")
                return ConsultStatus.SEM_SALDO, None, True, None
            if "Número de documento inválido" in combined_msg or "CPF inválido" in combined_msg:
                logging.warning(f"CPF inválido na planilha [{cpf}]: {combined_msg.strip()}")
                return ConsultStatus.CPF_INVALIDO, None, True, None
            if "Tente novamente" in combined_msg:
                logging.warning(f"Erro temporário da API para o CPF {cpf}: {combined_msg.strip()}")
                return ConsultStatus.RETRY, None, False, None
            
            logging.error(f"Falha de consulta (400) para CPF {cpf}: {combined_msg.strip()}")
            return None, None, False, f"HTTP 400: {detail or error_msg}"

        if response.status_code == 500:
            print(f"Erro encontrado: {error_msg}.")

            if error_msg in (
                'values() must be called with at least one value',
                "Saldo insuficiente, parcelas menores R$10,00.",
            ):
                logging.info(f"CPF {cpf}: Sem saldo")
                return ConsultStatus.SEM_SALDO, None, True, None

            if error_msg == "Cannot read properties of undefined (reading 'map')":
                print("J17 se encontra OFFLINE")
                logging.error("J17 OFFLINE (Erro map undefined)")
                return ConsultStatus.NAO_AUTORIZADO, None, True, None

            if any(x in error_msg for x in [
                "Falha ao buscar o saldo disponivel!",
                "Serviço indisponivel no momento, tente novamente mais tarde",
                "Excedido o limite de requisições (máximo de 1 por segundo).",
                "Limite de requisições excedido, tente novamente mais tarde",
            ]):
                return ConsultStatus.RETRY, None, False, None

            if any(x in error_msg for x in ["Instituição Fiduciária", "Cliente não autorizou", "Empty response"]):
                return ConsultStatus.NAO_AUTORIZADO, None, True, None

            if "Número de CPF inválido" in error_msg:
                logging.warning(f"CPF inválido: {cpf}")
                return ConsultStatus.CPF_INVALIDO, None, True, None

            if "Valor dos custos superior ao valor financiado" in error_msg or "Saldo insuficiente" in error_msg:
                logging.info(f"CPF {cpf}: sem saldo ({error_msg})")
                return ConsultStatus.SEM_SALDO, None, True, None

            logging.error(f"Erro 500 desconhecido para CPF {cpf}: {error_msg}")
            return None, None, False, f"HTTP 500: {error_msg}"

        reason = f"HTTP {response.status_code}: {response.text[:200]}"
        logging.error(f"Falha consulta CPF {cpf}. Status: {response.status_code}. Response: {response.text}")
        return None, None, False, reason

    except Exception as e:
        logging.error(f"Exceção ao processar resposta da API para CPF {cpf}: {e} | body bruto: {raw[:500]}", exc_info=True)
        return None, None, False, f"Erro ao processar resposta da API: {e}"


def consult_balance(
    session: requests.Session,
    token: str,
    cpf: str,
    averbador: str
) -> Tuple[Optional[List[Dict]], Optional[str], bool, Optional[str]]:
    """
    Consults the balance for a given CPF.
    Returns: (balance_periods, balance_id, success_flag, error_reason)
    error_reason is set only on FALHA_CONSULTA.
    """
    headers = {'Authorization': f'Bearer {token}'}
    data_consult = {
        "documentNumber": cpf,
        "provider": averbador
    }
    params_consult_get = {
        "search": cpf,
        "status": "success",
        "page": 1,
        "limit": 1
    }

    retries = 0
    while retries < MAX_RETRIES:
        try:
            response = session.post(
                URL_CONSULT_BALANCE,
                headers=headers,
                json=data_consult,
                timeout=TIMEOUT_SECONDS
            )
            response.raise_for_status()

            balance_data = response.json()
            if balance_data is None:
                # Polling loop for async result
                poll_retries = 0
                while poll_retries < 15:
                    response = session.get(
                        URL_CONSULT_BALANCE,
                        headers=headers,
                        params=params_consult_get,
                        timeout=TIMEOUT_SECONDS
                    )
                    response.raise_for_status()
                    balance_data = response.json()
                    if balance_data and balance_data.get('data'):
                        break
                    time.sleep(2)
                    poll_retries += 1

            if not balance_data:
                return None, None, False, "Resposta nula ou vazia após polling"

            data_list = balance_data.get('data', [])
            if not data_list:
                # Even after polling, data is empty
                return None, None, False, "Resposta sem dados (data vazio) após polling"

            first_record = data_list[0]
            balance_periods = first_record.get('periods', [])
            balance_id = first_record.get('id')

            print(f"balance_periods: {balance_periods}")
            if not isinstance(balance_periods, list):
                logging.error(f"Erro CPF {cpf}: balance_periods não é uma lista. Conteúdo: {balance_periods}")
                return None, None, False, "balance_periods inválido na resposta"

            new_balance_periods = [
                {"totalAmount": amount['amount'], "amount": amount['amount'], "dueDate": amount['dueDate']}
                for amount in balance_periods if isinstance(amount, dict)
            ]
            return new_balance_periods, balance_id, True, None

        except requests.RequestException as e:
            if getattr(e.response, 'status_code', None) == 429:
                print("Limite de requisições excedido, aguardando para tentar novamente...")
                logging.warning("Limite 429 excedido. Aguardando 2s...")
                time.sleep(2)
                retries += 1
                continue

            if e.response is not None:
                if e.response.status_code == 401:
                    logging.warning(f"CPF {cpf}: Token expirado (401). Solicitando renovação.")
                    return ConsultStatus.TOKEN_EXPIRADO, None, False, None

                result, b_id, finished, reason = _handle_consult_error(e.response, cpf)
                if result == ConsultStatus.RETRY:
                    print("Erro recuperável encontrado. Tentando novamente...")
                    time.sleep(0.4)
                    retries += 1
                    continue
                if finished:
                    return result, b_id, True, None
                return None, None, False, reason

            print(f"Falha ao consultar o BALANÇO. Sem resposta. Tentativa {retries + 1}/{MAX_RETRIES}")
            logging.warning(f"Falha consulta BALANÇO CPF {cpf}. Sem resposta. Tentativa {retries + 1}/{MAX_RETRIES}. Erro: {e}")
            time.sleep(2)
            retries += 1
            continue

        except Exception as e:
            print(f'Ocorreu um erro ao fazer a solicitação de PARCELAS: {e}')
            logging.error(f"Erro exceção ao solicitar PARCELAS CPF {cpf}: {e}", exc_info=True)
            retries += 1
            continue

    return None, None, False, "Número máximo de tentativas atingido"


def five_hour_stop() -> bool:
    """Asks the user if they want to stop the bot at 5 AM."""
    stop_five = input("Deseja parar o bot às 5h da manhã? (s/n): ")
    if stop_five.lower() == 's':
        print("O bot será pausado automaticamente às 5h da manhã.")
        return True
    print("O bot continuará executando normalmente.")
    return False


def select_tabela() -> Dict:
    """Interactive arrow-key menu to select simulation table."""
    opcoes = list(TABELAS.values())
    idx = 0

    KEY_UP    = b'H'
    KEY_DOWN  = b'P'
    KEY_ENTER = b'\r'

    RESET    = '\033[0m'
    SELECTED = '\033[1;97;45m'  # negrito, branco, fundo roxo
    NORMAL   = '\033[2;37m'     # cinza claro

    # Habilita sequências ANSI no terminal Windows
    os.system('')

    # Título + linha em branco + N opções = N+2 linhas no total
    TOTAL_LINES = len(opcoes) + 2

    def _render(current: int, first: bool = False):
        if not first:
            # Sobe exatamente o número de linhas que foram impressas
            sys.stdout.write(f'\033[{TOTAL_LINES}A')
        sys.stdout.write('\033[J')
        sys.stdout.write(' Escolha a tabela de simulação:\n\n')
        for i, op in enumerate(opcoes):
            if i == current:
                sys.stdout.write(f'  {SELECTED}  ▶  {op["nome"]}  {RESET}\n')
            else:
                sys.stdout.write(f'{NORMAL}     {op["nome"]}  {RESET}\n')
        sys.stdout.flush()

    _render(idx, first=True)

    while True:
        key = msvcrt.getch()

        if key == b'\xe0':
            key = msvcrt.getch()
            if key == KEY_UP:
                idx = (idx - 1) % len(opcoes)
                _render(idx)
            elif key == KEY_DOWN:
                idx = (idx + 1) % len(opcoes)
                _render(idx)
        elif key == KEY_ENTER:
            tabela = opcoes[idx]
            sys.stdout.write(f'\n  Tabela selecionada: {tabela["nome"]}\n\n')
            sys.stdout.flush()
            logging.info(f"Tabela selecionada: {tabela['nome']} (ID: {tabela['id']})")
            return tabela


def simulation(
    session: requests.Session,
    token: str,
    list_balance: List[Dict],
    cpf: str,
    balance_id: str,
    fees_id: str,
) -> Optional[float]:
    """Runs a simulation to get available balance."""
    headers = {'Authorization': f'Bearer {token}'}
    data_simulation = {
        "documentNumber": cpf,
        "isInsured": False,
        "simulationFeesId": fees_id,
        "targetAmount": 0,
        "provider": PROVIDER_BMS,
        "desiredInstallments": list_balance,
        "balanceId": balance_id
    }

    retries = 0
    while retries < MAX_RETRIES:
        try:
            response = session.post(
                URL_SIMULATION,
                headers=headers,
                json=data_simulation,
                timeout=TIMEOUT_SECONDS
            )

            if response.status_code >= 400:
                try:
                    body = response.json()
                    error_msg = body.get('error') or body.get('detail') or body.get('message')
                    if error_msg:
                        logging.error(f"CPF {cpf}: Erro na SIMULAÇÃO ({response.status_code}): {error_msg}")
                except ValueError:
                    logging.warning(f"CPF {cpf}: resposta não-JSON na SIMULAÇÃO (status {response.status_code}): {response.text[:200]}")
                    time.sleep(2)
                    retries += 1
                    continue
                response.raise_for_status()

            simulation_data = response.json()
            available_balance = simulation_data.get('availableBalance')
            
            if available_balance is not None:
                print(f'Saldo disponível para o CPF: {cpf}', available_balance)
                logging.info(f"Saldo disponível CPF {cpf}: {available_balance}")
                return available_balance
            
            print(f'Simulação não retornou saldo disponível para o CPF: {cpf}')
            return None

        except requests.RequestException as e:
            if getattr(e.response, 'status_code', None) == 429:
                print("Limite de requisições excedido na simulação, aguardando...")
                time.sleep(2)
                retries += 1
                continue
            
            if getattr(e.response, 'status_code', 0) >= 500:
                try:
                    error_body = e.response.json()
                    detail = error_body.get('error') or error_body.get('detail') or error_body.get('message') or ""
                    print(f"Erro no servidor durante simulação: {e.response.status_code}. {detail}. Tentando novamente...")
                except Exception:
                    print(f"Erro no servidor durante simulação: {e.response.status_code}. Tentando novamente...")
                
                time.sleep(2)
                retries += 1
                continue

            if getattr(e, 'response', None) is None:
                print(f"Falha ao simular SALDO. Sem resposta. Tentativa {retries + 1}/{MAX_RETRIES}")
                logging.warning(f"Erro solicitação SALDO CPF {cpf}. Sem resposta. Tentativa {retries + 1}/{MAX_RETRIES}. Erro: {e}")
                time.sleep(2)
                retries += 1
                continue

            print(f'Erro fatal ao fazer a solicitação de SALDO: {e}')
            logging.error(f"Erro fatal solicitação SALDO CPF {cpf}: {e}")
            break

    return None


def process_row(
    row: Tuple,
    token: str,
    session: requests.Session,
    fees_id: str,
) -> Tuple[str, Union[str, float, ConsultStatus], Optional[str]]:
    """Processes a single row (CPF).
    Returns: (cpf, result_value, error_reason)
    error_reason is only set when result is FALHA_CONSULTA.
    May return ConsultStatus.TOKEN_EXPIRADO to signal renewal.
    """
    cpf = str(row[0])

    print(f"Consultando CPF:{cpf} com '{PROVIDER_BMS}'...")
    logging.info(f"Consultando CPF: {cpf}")
    time.sleep(2)

    balance, balance_id, success, reason = consult_balance(session, token, cpf, PROVIDER_BMS)

    if balance == ConsultStatus.TOKEN_EXPIRADO:
        return cpf, ConsultStatus.TOKEN_EXPIRADO, None

    if success and balance in (ConsultStatus.NAO_AUTORIZADO, ConsultStatus.CPF_INVALIDO):
        return cpf, balance.value, None

    if success and balance not in (ConsultStatus.SEM_SALDO, None):
        if len(balance) < 2:
            logging.info(f"CPF {cpf}: Sem saldo (menos de 2 parcelas disponíveis)")
            return cpf, ConsultStatus.SEM_SALDO.value, None
            
        sim = simulation(session, token, balance, cpf, balance_id, fees_id)
        if sim is not None:
            return cpf, sim, None
        return cpf, ConsultStatus.FALHA_CONSULTA.value, "Simulação não retornou saldo disponível"

    if not success:
        return cpf, ConsultStatus.FALHA_CONSULTA.value, reason

    return cpf, ConsultStatus.SEM_SALDO.value, None


def bot_auth(user_id: str, token: str) -> Optional[Dict]:
    """Authenticates the bot user."""
    url = f'{URL_BOT_AUTH_BASE}{user_id}/'
    headers = {'Authorization': f'Token {token}'}

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f'Erro na autenticação do bot: {e}')
        logging.critical(f"Erro na autenticação do bot: {e}")
        return None


def _get_or_prompt_env(key: str, prompt: str, secret: bool = False) -> str:
    """Returns env var value, prompting and persisting if missing."""
    value = os.getenv(key, '')
    if not value:
        value = getpass.getpass(prompt) if secret else input(prompt)
        os.environ[key] = value
        set_key(ENV_FILE_PATH, key, value)
    return value


def update_env() -> Tuple[str, str, str, str]:
    """Updates/Loads environment variables."""
    load_dotenv(ENV_FILE_PATH)
    user_id = _get_or_prompt_env('USER_ID', "Digite o seu ID: ", secret=True)
    user_name = _get_or_prompt_env('USER_NAME', "Usuário V8: ")
    user_password = _get_or_prompt_env('USER_PASSWORD', "Senha V8: ")
    user_token = _get_or_prompt_env('USER_TOKEN', "Token de Auth: ", secret=True)
    return user_id, user_name, user_password, user_token


def load_counter() -> int:
    """Loads the processed counter from file."""
    try:
        with open(CONTADOR_FILE, 'r') as f:
            return int(f.read().strip())
    except (FileNotFoundError, ValueError):
        return 0
    except Exception as e:
        print(f"Erro ao ler o arquivo de contador: {e}")
        return 0


def save_counter(counter: int):
    """Saves the processed counter to file."""
    try:
        with open(CONTADOR_FILE, 'w') as f:
            f.write(str(counter))
    except Exception as e:
        print(f"Erro ao salvar contador: {e}")


def load_results() -> Optional[Dict]:
    """Loads the current result payload from JSON file."""
    try:
        if os.path.exists(RESULT_JSON_FILE):
            with open(RESULT_JSON_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Erro ao carregar resultados JSON: {e}")
        logging.error(f"Erro ao carregar resultados JSON: {e}")
    return None


def save_results(payload: Dict):
    """Saves the current result payload (meta + cpfs) to JSON file."""
    try:
        if not os.path.exists(RESULT_JSON_DIR):
            os.makedirs(RESULT_JSON_DIR)
        with open(RESULT_JSON_FILE, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Erro ao salvar resultados: {e}")
        logging.error(f"Erro ao salvar resultados JSON: {e}")


def main():
    """Main execution entry point."""
    user_id, v8_username, v8_password, api_token = update_env()
    
    # Setup Logger
    try:
        log_file = setup_logger()
        print(f"Logs sendo salvos em: {log_file}")
    except Exception as e:
        print(f"Erro ao configurar sistema de logs: {e}")

    stop_at_five = five_hour_stop()

    try:
        user_info = bot_auth(user_id, api_token)
        if not user_info:
            input('Pressione ENTER para sair.')
            return
    except Exception as e:
        err_msg = f"Autenticação fora do ar: {e}"
        print(err_msg)
        logging.critical(err_msg)
        input('Pressione ENTER para sair.')
        return

    user_cpf = input("Digite o seu CPF (sem ponto e sem traço): ")

    if user_info.get('cpf') != user_cpf:
        print("CPF não pertence a esse usuário, contate o administrador para mais detalhes.")
        input('Pressione ENTER para sair.')
        return

    if not user_info.get('active'):
        print("Usuário inativo, contate o administrador para mais detalhes.")
        logging.warning(f"Tentativa de login usuário inativo: {user_info.get('username')}")
        input('Pressione ENTER para sair.')
        return

    if not os.path.exists(BASE_DIR):
        os.makedirs(BASE_DIR)
        print(f"A pasta '{BASE_DIR}' foi criada. Coloque o arquivo XLSX dentro dela e execute novamente.")
        input('Pressione ENTER para sair.')
        return

    base_files = [f for f in os.listdir(BASE_DIR) if not f.startswith('~$')]
    if len(base_files) == 0:
        print(f"A pasta '{BASE_DIR}' está vazia. Coloque o arquivo XLSX dentro dela e execute novamente.")
        input('Pressione ENTER para sair.')
        return
    if len(base_files) != 1:
        print(f"A pasta '{BASE_DIR}' deve conter exatamente um arquivo XLSX. Foram encontrados {len(base_files)} arquivos.")
        input('Pressione ENTER para sair.')
        return

    arquivo_xlsx = os.path.join(BASE_DIR, base_files[0])
    
    # machine_id = str(uuid.uuid1()) # Unused in original code? Keeping it out to clean up unless needed.

    print(f"Olá, {user_info.get('name')}! Iniciando consultas.")
    logging.info(f"Iniciando sessão para {user_info.get('name')} (ID: {user_info.get('id')})")

    tabela = select_tabela()
    fees_id = tabela["id"]

    token = get_token(v8_username, v8_password)
    if token is None:
        print("Não foi possível obter o token de autenticação.")
        logging.critical("Falha ao obter token de autenticação.")
        return

    contador = load_counter()

    try:
        wb = openpyxl.load_workbook(arquivo_xlsx)
        sheet = wb.active
        total_rows = sheet.max_row - 1  # descontando cabeçalho

        inicio = datetime.now()
        
        existing_payload = load_results() if contador > 0 else None
        
        if not existing_payload and contador > 0:
            print("Histórico (result.json) não encontrado. Reiniciando do começo (contador = 0).")
            contador = 0
            save_counter(0)
        
        if existing_payload and "meta" in existing_payload and "cpfs" in existing_payload:
            payload = existing_payload
        else:
            payload = {
                "meta": {
                    "tabela": base_files[0],
                    "tabela_simulacao": tabela["nome"],
                    "operador": user_info.get('name'),
                    "inicio": inicio.strftime("%Y-%m-%d %H:%M:%S"),
                    "fim": None,
                    "total_cpfs": total_rows,
                    "processados": contador,
                    "com_saldo": 0,
                    "sem_saldo": 0,
                    "nao_autorizado": 0,
                    "cpf_invalido": 0,
                    "operacao_em_andamento": 0,
                    "falha_consulta": 0,
                },
                "cpfs": {}
            }
        save_results(payload)

        session = create_session()

        rows_iter = sheet.iter_rows(min_row=contador + 2, max_col=1, values_only=True)

        for idx, row in enumerate(rows_iter, start=contador + 2):

            # 5 AM Check
            now = datetime.now()
            if stop_at_five and now.hour == 5:
                print("\n⚠️ O bot foi pausado automaticamente às 05:00 da manhã. ⚠️")
                logging.info("Bot pausado automaticamente (05:00).")
                input("Pressione ENTER para encerrar.")
                return

            cpf, result, reason = process_row(row, token, session, fees_id)

            if result == ConsultStatus.TOKEN_EXPIRADO:
                print("Token expirado. Renovando...")
                logging.warning("Token expirado. Tentando renovar...")
                token = get_token(v8_username, v8_password)
                if token is None:
                    print("Não foi possível renovar o token. Encerrando.")
                    logging.critical("Falha ao renovar token. Encerrando.")
                    return
                print("Token renovado. Retentando CPF...")
                cpf, result, reason = process_row(row, token, session, fees_id)

            result_value = result if not isinstance(result, ConsultStatus) else result.value
            consulta_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            entry: Dict[str, Any] = {'Valor': result_value, 'Data e Hora': consulta_date}
            if reason:
                entry['Motivo'] = reason
            payload["cpfs"][cpf] = entry

            # Atualiza contadores de resumo
            meta = payload["meta"]
            if isinstance(result, (int, float)):
                meta["com_saldo"] += 1
            elif result_value == ConsultStatus.SEM_SALDO.value:
                meta["sem_saldo"] += 1
            elif result_value == ConsultStatus.NAO_AUTORIZADO.value:
                meta["nao_autorizado"] += 1
            elif result_value == ConsultStatus.CPF_INVALIDO.value:
                meta["cpf_invalido"] += 1
            elif result_value == ConsultStatus.OPERACAO_EM_ANDAMENTO.value:
                meta["operacao_em_andamento"] = meta.get("operacao_em_andamento", 0) + 1
            else:
                meta["falha_consulta"] += 1

            contador += 1
            meta["processados"] = contador

            save_counter(contador)
            save_results(payload)

            print(f"CPF: {cpf} = Resultado: {result_value}")

        payload["meta"]["fim"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_results(payload)
        save_counter(0)
        print("Processamento concluído com sucesso.")
        logging.info("Processamento concluído com sucesso.")

    except KeyboardInterrupt:
        print("\nOperação interrompida pelo usuário.")
        logging.warning("Operação interrompida pelo usuário (KeyboardInterrupt).")
    except FileNotFoundError:
        print("Arquivo XLSX não encontrado.")
        logging.error("Arquivo XLSX não encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro durante o processamento: {e}")
        logging.error(f"Erro fatal durante processamento main loop: {e}", exc_info=True)


if __name__ == '__main__':
    main()
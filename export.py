"""
Script to export results from JSON to Excel.
"""

import json
import os
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Constants
RESULT_DIR_NAME = 'result_json'
RESULT_FILENAME = 'result.json'
RESULT_PATH = os.path.join(RESULT_DIR_NAME, RESULT_FILENAME)
OUTPUT_DIR = 'output'
DATE_FORMAT_FILE = "%d-%m"
FONT_NAME = 'Calibri'

# --- Cores ---
# Resultados
COR_HEADER_RESULT  = 'C8A2C8'  # roxo claro
COR_COM_SALDO      = 'C6EFCE'  # verde claro
COR_SEM_SALDO      = 'FCE4D6'  # laranja claro
COR_NAO_AUTORIZADO = 'FFEB9C'  # amarelo claro
COR_FALHA          = 'FFC7CE'  # vermelho claro
COR_NEUTRO         = 'FFFFFF'  # branco (CPF inválido e outros)

# Resumo
COR_HEADER_RESUMO  = '4472C4'  # azul neutro (texto branco)
COR_CORPO_RESUMO   = 'DDEEFF'  # azul clarinho


def _font(bold=False, color='000000', size=11):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)


def _fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)


def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_cell(cell, value, bold=False, font_color='000000', bg_color=None,
                align='left', border=True):
    cell.value = value
    cell.font = _font(bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if bg_color:
        cell.fill = _fill(bg_color)
    if border:
        cell.border = _thin_border()


def _row_color_for(valor) -> str:
    if isinstance(valor, (int, float)):
        return COR_COM_SALDO
    v = str(valor)
    if v == 'SEM SALDO':
        return COR_SEM_SALDO
    if v == 'NAO AUTORIZADO':
        return COR_NAO_AUTORIZADO
    if v == 'FALHA CONSULTA':
        return COR_FALHA
    return COR_NEUTRO


def export_to_excel(payload: Dict) -> None:
    meta = payload.get("meta", {})
    cpfs_data = payload.get("cpfs", {})

    if not cpfs_data:
        print("Nenhum dado para exportar.")
        return

    # Monta linhas dos resultados
    result_rows: List[Dict] = []
    has_motivo = False
    for cpf, info in cpfs_data.items():
        motivo = info.get('Motivo')
        if motivo:
            has_motivo = True
        result_rows.append({
            'CPF': cpf,
            'Valor': info.get('Valor'),
            'Motivo': motivo,
            'Data e Hora': info.get('Data e Hora'),
        })

    resumo_rows = [
        ("Tabela processada",       meta.get("tabela", "-")),
        ("Tabela de simulação",     meta.get("tabela_simulacao", "-")),
        ("Operador",                meta.get("operador", "-")),
        ("Início",                  meta.get("inicio", "-")),
        ("Fim",                     meta.get("fim", "-")),
        ("Total CPFs na tabela",    meta.get("total_cpfs", "-")),
        ("CPFs processados",        meta.get("processados", "-")),
        ("Com saldo",               meta.get("com_saldo", 0)),
        ("Sem saldo",               meta.get("sem_saldo", 0)),
        ("Não autorizado",          meta.get("nao_autorizado", 0)),
        ("CPF inválido",            meta.get("cpf_invalido", 0)),
        ("Falha de consulta",       meta.get("falha_consulta", 0)),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.sheet_view.showGridLines = False

    # ── Colunas de resultados ──────────────────────────────────────────
    # Col A: CPF, B: Valor, C: Motivo (opcional), D: Data e Hora
    # Gap: col E vazia
    # Col F: Campo resumo, G: Valor resumo
    if not has_motivo:
        # A=CPF, B=Valor, C=Data e Hora  |  D=gap  |  E=Campo, F=Valor
        col_res_f = 'E'
        col_res_v = 'F'
    else:
        # A=CPF, B=Valor, C=Motivo, D=Data e Hora  |  E=gap  |  F=Campo, G=Valor
        col_res_f = 'F'
        col_res_v = 'G'

    # Larguras
    ws.column_dimensions['A'].width = 16   # CPF
    ws.column_dimensions['B'].width = 16   # Valor
    if has_motivo:
        ws.column_dimensions['C'].width = 52   # Motivo
        ws.column_dimensions['D'].width = 20   # Data e Hora
        ws.column_dimensions['E'].width = 2    # gap
        ws.column_dimensions['F'].width = 26   # Campo resumo
        ws.column_dimensions['G'].width = 26   # Valor resumo
    else:
        ws.column_dimensions['C'].width = 20   # Data e Hora
        ws.column_dimensions['D'].width = 2    # gap
        ws.column_dimensions['E'].width = 26   # Campo resumo
        ws.column_dimensions['F'].width = 26   # Valor resumo

    # ── Header resultados ──────────────────────────────────────────────
    ROW_HEADER = 1
    ws.row_dimensions[ROW_HEADER].height = 22
    _apply_cell(ws['A1'], 'CPF',         bold=True, font_color='FFFFFF', bg_color=COR_HEADER_RESULT)
    _apply_cell(ws['B1'], 'Valor',       bold=True, font_color='FFFFFF', bg_color=COR_HEADER_RESULT)
    if has_motivo:
        _apply_cell(ws['C1'], 'Motivo da Falha', bold=True, font_color='FFFFFF', bg_color=COR_HEADER_RESULT)
        _apply_cell(ws['D1'], 'Data e Hora',     bold=True, font_color='FFFFFF', bg_color=COR_HEADER_RESULT)
    else:
        _apply_cell(ws['C1'], 'Data e Hora', bold=True, font_color='FFFFFF', bg_color=COR_HEADER_RESULT)

    # ── Corpo resultados ───────────────────────────────────────────────
    for i, row in enumerate(result_rows, start=2):
        ws.row_dimensions[i].height = 16
        bg = _row_color_for(row['Valor'])
        valor = row['Valor']
        # Formata valor numérico com R$
        display_valor = f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if isinstance(valor, (int, float)) else valor
        _apply_cell(ws[f'A{i}'], row['CPF'],       bg_color=bg)
        _apply_cell(ws[f'B{i}'], display_valor,    bg_color=bg, align='center')
        if has_motivo:
            _apply_cell(ws[f'C{i}'], row['Motivo'] or '', bg_color=bg)
            _apply_cell(ws[f'D{i}'], row['Data e Hora'],  bg_color=bg, align='center')
        else:
            _apply_cell(ws[f'C{i}'], row['Data e Hora'],  bg_color=bg, align='center')

    # ── Header resumo ──────────────────────────────────────────────────
    _apply_cell(ws[f'{col_res_f}1'], 'Campo', bold=True, font_color='FFFFFF', bg_color=COR_HEADER_RESUMO)
    _apply_cell(ws[f'{col_res_v}1'], 'Valor', bold=True, font_color='FFFFFF', bg_color=COR_HEADER_RESUMO)

    # ── Corpo resumo ───────────────────────────────────────────────────
    for i, (campo, valor) in enumerate(resumo_rows, start=2):
        ws.row_dimensions[i].height = 16
        _apply_cell(ws[f'{col_res_f}{i}'], campo, bold=False, bg_color=COR_CORPO_RESUMO)
        _apply_cell(ws[f'{col_res_v}{i}'], valor, bold=False, bg_color=COR_CORPO_RESUMO, align='center')

    # ── Freeze header ──────────────────────────────────────────────────
    ws.freeze_panes = 'A2'

    # ── Salva ──────────────────────────────────────────────────────────
    now = datetime.now()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    name_file = os.path.join(OUTPUT_DIR, f'Saldos_{now.strftime(DATE_FORMAT_FILE)}.xlsx')

    try:
        wb.save(name_file)
        print(f"Arquivo exportado com sucesso: {name_file}")
    except Exception as e:
        print(f"Erro ao salvar Excel: {e}")


def load_json_data(filepath: str) -> Optional[Dict]:
    if not os.path.exists(filepath):
        print(f"Arquivo não encontrado: {filepath}")
        return None
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            return json.load(file)
    except Exception as e:
        print(f"Erro ao ler arquivo JSON: {e}")
        return None


def cleanup_file(filepath: str) -> None:
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print("Arquivo temporário removido.")
        if os.path.exists('contador.txt'):
            os.remove('contador.txt')
            print("Contador resetado.")
    except Exception as e:
        print(f"Erro ao remover arquivo: {e}")


def main():
    print("Iniciando exportação...")
    payload = load_json_data(RESULT_PATH)

    if payload:
        if "cpfs" not in payload:
            payload = {"meta": {}, "cpfs": payload}
        export_to_excel(payload)
        cleanup_file(RESULT_PATH)
    else:
        print("Processo finalizado sem exportação.")

    input("\nPressione ENTER para sair.")


if __name__ == '__main__':
    main()
